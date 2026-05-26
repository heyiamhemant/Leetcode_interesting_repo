"""Parse bakeries-india-inventory.md and emit an .xlsx with one row per item.

Schema: State | City | Bakery | Area | Category | Item | Price | Source URL(s)

Notes
- Cities map to Indian states; "Delhi" is treated as a UT/state of its own.
- Bullet lines containing an em dash followed by a URL are item rows. A
  bullet line without a URL is still captured as an item (price/url empty).
- Lines starting with "> " are blockquote notes from the markdown (Zomato
  section roll-ups, etc.) and are captured as a single row with category
  "Section roll-up (app-only SKUs)".
- "**Sources:**" lines under a bakery header are captured as the bakery's
  source list and shown on every item row from that bakery whose own URL
  is missing.
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MD_PATH = Path("/Users/macandcode/Documents/Leetcode_interesting_repo/bakeries-india-inventory.md")
XLSX_PATH = Path("/Users/macandcode/Documents/Leetcode_interesting_repo/bakeries-india-inventory.xlsx")

CITY_STATE = {
    "Mumbai": "Maharashtra",
    "Pune": "Maharashtra",
    "Nagpur": "Maharashtra",
    "Delhi": "Delhi",
    "Kolkata": "West Bengal",
    "Siliguri": "West Bengal",
    "Darjeeling": "West Bengal",
    "Chennai": "Tamil Nadu",
    "Madurai": "Tamil Nadu",
    "Bangalore (Bengaluru)": "Karnataka",
    "Bangalore": "Karnataka",
}

URL_RE = re.compile(r"https?://\S+")
PRICE_RE = re.compile(r"₹[\d,]+(?:\.\d+)?(?:\s*/\s*₹?[\d,]+(?:\.\d+)?)*|From\s+₹[\d,]+(?:\.\d+)?", re.IGNORECASE)
BOLD_NAME_RE = re.compile(r"^\*\*(.+?)\*\*")
CITY_RE = re.compile(r"^#\s+(.+?)\s*$")
BAKERY_RE = re.compile(r"^##\s+(.+?)\s*$")
CATEGORY_RE = re.compile(r"^###\s+(.+?)\s*$")
SOURCES_RE = re.compile(r"^\*\*Sources:\*\*\s*(.+)$")
STATUS_RE = re.compile(r"^\*\*Status:\*\*\s*(.+)$")


def split_urls(text: str) -> list[str]:
    return [u.rstrip(".,;)") for u in URL_RE.findall(text)]


def parse_bullet(line: str) -> tuple[str, str, list[str]] | None:
    """Return (item_name, price, urls) from a bullet line, or None."""
    if not line.startswith("- "):
        return None
    body = line[2:].strip()
    m = BOLD_NAME_RE.match(body)
    if m:
        name = m.group(1).strip()
        rest = body[m.end():]
    else:
        # Plain bullet (e.g. inside an aggregated section roll-up)
        name = body.split(" — ", 1)[0].strip()
        rest = body[len(name):]
    urls = split_urls(rest)
    price_match = PRICE_RE.search(rest.split(" — ", 1)[0]) or PRICE_RE.search(rest)
    price = price_match.group(0).strip() if price_match else ""
    return name, price, urls


def parse_blockquote_items(text: str) -> list[str]:
    """Pull bracketed counts/items from a Zomato roll-up blockquote string."""
    inner = re.sub(r"^>\s*", "", text).strip().strip("*")
    return [chunk.strip() for chunk in re.split(r";|\.", inner) if chunk.strip()]


def main() -> None:
    md = MD_PATH.read_text(encoding="utf-8")
    lines = md.splitlines()

    rows: list[dict] = []
    city: str | None = None
    bakery: str | None = None
    area: str | None = None
    category: str | None = None
    bakery_sources: list[str] = []
    status_note: str | None = None

    in_accuracy_notes = False

    for raw in lines:
        line = raw.rstrip()

        if line.startswith("# "):
            new_city = CITY_RE.match(line).group(1)
            if new_city.lower().startswith("bakeries of india"):
                continue
            city = new_city
            bakery = None
            category = None
            bakery_sources = []
            status_note = None
            continue

        if line.startswith("## "):
            m = BAKERY_RE.match(line)
            header = m.group(1)
            # Header is usually "Bakery Name — Area / address"
            if " — " in header:
                bakery, area = [p.strip() for p in header.split(" — ", 1)]
            else:
                bakery, area = header.strip(), ""
            category = None
            bakery_sources = []
            status_note = None
            in_accuracy_notes = bakery.lower().startswith("accuracy notes")
            continue

        if in_accuracy_notes:
            continue

        if line.startswith("### "):
            category = CATEGORY_RE.match(line).group(1)
            continue

        sm = SOURCES_RE.match(line)
        if sm:
            bakery_sources = split_urls(sm.group(1))
            continue

        st = STATUS_RE.match(line)
        if st:
            status_note = st.group(1).strip()
            # Emit a status row so the sparse bakeries show up in the sheet
            if city and bakery:
                rows.append({
                    "state": CITY_STATE.get(city, ""),
                    "city": city,
                    "bakery": bakery,
                    "area": area or "",
                    "category": "(status)",
                    "item": status_note,
                    "price": "",
                    "source": "; ".join(bakery_sources),
                })
            continue

        if line.startswith("> "):
            if city and bakery:
                rows.append({
                    "state": CITY_STATE.get(city, ""),
                    "city": city,
                    "bakery": bakery,
                    "area": area or "",
                    "category": "Section roll-up (app-only SKUs)",
                    "item": re.sub(r"^>\s*", "", line).strip(),
                    "price": "",
                    "source": "; ".join(split_urls(line)) or "; ".join(bakery_sources),
                })
            continue

        bullet = parse_bullet(line)
        if bullet and city and bakery:
            name, price, urls = bullet
            rows.append({
                "state": CITY_STATE.get(city, ""),
                "city": city,
                "bakery": bakery,
                "area": area or "",
                "category": category or "(uncategorised)",
                "item": name,
                "price": price,
                "source": "; ".join(urls) if urls else "; ".join(bakery_sources),
            })

    # Build the workbook ---------------------------------------------------
    wb = Workbook()

    # Sheet 1 — flat master list
    master = wb.active
    master.title = "All items"
    headers = ["State", "City", "Bakery", "Area", "Category", "Item", "Price", "Source URL(s)"]
    master.append(headers)
    for r in rows:
        master.append([
            r["state"], r["city"], r["bakery"], r["area"],
            r["category"], r["item"], r["price"], r["source"],
        ])

    # Sheet 2 — per-bakery item counts
    summary = wb.create_sheet("Summary by bakery")
    summary.append(["State", "City", "Bakery", "Area", "Items documented", "Has 'Status' note"])
    bakery_index: dict[tuple[str, str, str], dict] = {}
    for r in rows:
        key = (r["state"], r["city"], r["bakery"])
        b = bakery_index.setdefault(key, {"area": r["area"], "items": 0, "status": False})
        if r["category"] == "(status)":
            b["status"] = True
        else:
            b["items"] += 1
    for (state, city, bakery), info in bakery_index.items():
        summary.append([state, city, bakery, info["area"], info["items"], "Yes" if info["status"] else ""])

    # Sheet 3 — counts by state / city
    counts = wb.create_sheet("Counts by city")
    counts.append(["State", "City", "Bakeries", "Total items documented"])
    city_index: dict[tuple[str, str], dict] = {}
    for (state, city, _bakery), info in bakery_index.items():
        c = city_index.setdefault((state, city), {"bakeries": 0, "items": 0})
        c["bakeries"] += 1
        c["items"] += info["items"]
    for (state, city), info in sorted(city_index.items()):
        counts.append([state, city, info["bakeries"], info["items"]])

    # Sheet 4 — counts by state
    state_counts = wb.create_sheet("Counts by state")
    state_counts.append(["State", "Cities", "Bakeries", "Total items documented"])
    st_idx: dict[str, dict] = {}
    for (state, city), info in city_index.items():
        s = st_idx.setdefault(state, {"cities": set(), "bakeries": 0, "items": 0})
        s["cities"].add(city)
        s["bakeries"] += info["bakeries"]
        s["items"] += info["items"]
    for state, info in sorted(st_idx.items()):
        state_counts.append([state, len(info["cities"]), info["bakeries"], info["items"]])

    # Styling --------------------------------------------------------------
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E78")
    wrap = Alignment(wrap_text=True, vertical="top")

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = head_font
            cell.fill = head_fill
        sheet.freeze_panes = "A2"
        for col_idx, col_cells in enumerate(sheet.columns, start=1):
            max_len = 0
            for c in col_cells:
                if c.value is None:
                    continue
                max_len = max(max_len, min(len(str(c.value)), 80))
                c.alignment = wrap
            sheet.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)

    # Make the master 'Item' column wide and source narrower-but-wrapped
    master.column_dimensions["F"].width = 60
    master.column_dimensions["H"].width = 80

    # Auto-filter on the master sheet
    master.auto_filter.ref = master.dimensions

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH} ({len(rows)} item rows, {len(bakery_index)} bakeries, {len(st_idx)} states)")


if __name__ == "__main__":
    main()
